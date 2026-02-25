"""
apps/data_import/views.py

Upload and import endpoints for Excel files.

Endpoints:
    POST   /api/import/upload/        — Upload and process a single Excel file
    GET    /api/import/logs/          — List import history for the company
    GET    /api/import/logs/{id}/     — Detail of a single import log
    DELETE /api/import/logs/{id}/     — Remove an import log record
    GET    /api/import/detect/        — Detect file type without importing (preview)
"""

import logging
from datetime import datetime, timezone

from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import ImportLog
from .serializers import ImportLogSerializer, ImportUploadSerializer
from .parsers.excel_parser import parse_excel_file, detect_file_type
import openpyxl

logger = logging.getLogger(__name__)


class ExcelUploadView(APIView):
    """
    POST /api/import/upload/

    Accepts a multipart form upload with an Excel file.
    Automatically detects the file type and triggers the appropriate parser.
    Creates an ImportLog entry and returns the result.

    Access rules:
        - Admin   : can import any file type
        - Manager : can import any file type for their company
        - Agent   : can import movements and inventory only
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    AGENT_ALLOWED_TYPES = {"movements", "inventory"}

    def post(self, request):
        serializer = ImportUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        file_obj = serializer.validated_data["file"]
        file_type_override = serializer.validated_data.get("file_type")
        snapshot_date = serializer.validated_data.get("snapshot_date")
        report_date = serializer.validated_data.get("report_date")

        company = request.user.company
        if not company:
            return Response(
                {"error": "Your account is not linked to a company. Contact your administrator."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Create import log (pending)
        log = ImportLog.objects.create(
            company=company,
            imported_by=request.user,
            file_type=file_type_override or "movements",  # Placeholder until detected
            original_filename=file_obj.name,
            status=ImportLog.ImportStatus.PROCESSING,
            import_context={
                "snapshot_date": str(snapshot_date) if snapshot_date else None,
                "report_date": str(report_date) if report_date else None,
            },
        )

        try:
            extra_context = {
                "snapshot_date": snapshot_date,
                "report_date": report_date,
            }

            result = parse_excel_file(
                file_obj=file_obj,
                filename=file_obj.name,
                company=company,
                file_type=file_type_override,
                extra_context=extra_context,
            )

            # Enforce agent file type restrictions
            detected_type = result["file_type"]
            if (
                request.user.is_agent
                and detected_type not in self.AGENT_ALLOWED_TYPES
            ):
                log.status = ImportLog.ImportStatus.FAILED
                log.error_details = [{"error": "Agents can only import movements and inventory files."}]
                log.completed_at = datetime.now(tz=timezone.utc)
                log.save()
                return Response(
                    {
                        "error": (
                            f"As an agent, you are not allowed to import "
                            f"'{detected_type}' files. Allowed types: "
                            f"{', '.join(self.AGENT_ALLOWED_TYPES)}."
                        )
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )

            # Update log with results
            has_errors = len(result.get("errors", [])) > 0
            log.file_type = detected_type
            log.row_count = result["total"]
            log.success_count = result["created"] + result.get("updated", 0)
            log.error_count = result["error_count"] if "error_count" in result else len(result.get("errors", []))
            log.error_details = result.get("errors", [])[:100]  # Cap at 100 stored errors
            log.status = (
                ImportLog.ImportStatus.PARTIAL
                if has_errors and log.success_count > 0
                else ImportLog.ImportStatus.FAILED
                if has_errors and log.success_count == 0
                else ImportLog.ImportStatus.SUCCESS
            )
            log.completed_at = datetime.now(tz=timezone.utc)
            log.save()

            logger.info(
                f"[ExcelUploadView] Import complete: '{file_obj.name}' "
                f"({detected_type}) for company '{company.name}' "
                f"— {log.success_count}/{log.row_count} rows, "
                f"{log.error_count} errors."
            )

            return Response(
                {
                    "message": self._build_message(log),
                    "import_log": ImportLogSerializer(log).data,
                    "result": {
                        "file_type": detected_type,
                        "total_rows": result["total"],
                        "created": result["created"],
                        "updated": result.get("updated", 0),
                        "errors": result.get("errors", [])[:20],  # Return first 20 errors
                    },
                },
                status=status.HTTP_201_CREATED,
            )

        except ValueError as e:
            log.status = ImportLog.ImportStatus.FAILED
            log.error_details = [{"error": str(e)}]
            log.completed_at = datetime.now(tz=timezone.utc)
            log.save()
            return Response(
                {"error": str(e), "import_log_id": str(log.id)},
                status=status.HTTP_422_UNPROCESSABLE_ENTITY,
            )

        except Exception as e:
            logger.exception(f"[ExcelUploadView] Unexpected error: {e}")
            log.status = ImportLog.ImportStatus.FAILED
            log.error_details = [{"error": f"Internal error: {str(e)}"}]
            log.completed_at = datetime.now(tz=timezone.utc)
            log.save()
            return Response(
                {"error": "An unexpected error occurred during import. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @staticmethod
    def _build_message(log: ImportLog) -> str:
        if log.status == ImportLog.ImportStatus.SUCCESS:
            return (
                f"Import successful: {log.success_count} records imported "
                f"from '{log.original_filename}'."
            )
        if log.status == ImportLog.ImportStatus.PARTIAL:
            return (
                f"Partial import: {log.success_count} records imported, "
                f"{log.error_count} rows failed."
            )
        return f"Import failed: {log.error_count} errors encountered."


class DetectFileTypeView(APIView):
    """
    POST /api/import/detect/

    Accepts a file upload and returns the detected file type + preview
    of the first 5 rows WITHOUT persisting any data.
    Useful for UI validation before actual import.
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response(
                {"error": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ext = file_obj.name.rsplit(".", 1)[-1].lower()
        if ext not in ("xlsx", "xls"):
            return Response(
                {"error": "Only .xlsx and .xls files are accepted."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(max_row=6, values_only=True))
            wb.close()
        except Exception as e:
            return Response(
                {"error": f"Cannot read file: {e}"},
                status=status.HTTP_422_UNPROCESSABLE_ENTITY,
            )

        header_row = rows[0] if rows else ()
        detected_type = detect_file_type(file_obj.name, header_row)

        preview = []
        if rows:
            headers = [str(h or "") for h in rows[0]]
            for row in rows[1:6]:
                preview.append(dict(zip(headers, [str(v or "") for v in row])))

        return Response({
            "filename": file_obj.name,
            "detected_file_type": detected_type,
            "headers": [str(h or "") for h in header_row],
            "preview_rows": preview,
            "total_rows_estimate": ws.max_row - 1 if hasattr(ws, "max_row") else None,
        })


class ImportLogListView(APIView):
    """
    GET /api/import/logs/
    Returns import history for the authenticated user's company.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = ImportLog.objects.filter(
            company=request.user.company
        ).select_related("imported_by").order_by("-started_at")

        file_type = request.query_params.get("file_type")
        if file_type:
            qs = qs.filter(file_type=file_type)

        status_filter = request.query_params.get("status")
        if status_filter:
            qs = qs.filter(status=status_filter)

        serializer = ImportLogSerializer(qs[:100], many=True)
        return Response({
            "count": qs.count(),
            "logs": serializer.data,
        })


class ImportLogDetailView(APIView):
    """
    GET    /api/import/logs/{id}/  → Import log details with full error list
    DELETE /api/import/logs/{id}/  → Delete an import log record
    """
    permission_classes = [IsAuthenticated]

    def _get_log(self, log_id, company):
        try:
            return ImportLog.objects.get(id=log_id, company=company)
        except ImportLog.DoesNotExist:
            return None

    def get(self, request, log_id):
        log = self._get_log(log_id, request.user.company)
        if not log:
            return Response({"error": "Import log not found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(ImportLogSerializer(log).data)

    def delete(self, request, log_id):
        log = self._get_log(log_id, request.user.company)
        if not log:
            return Response({"error": "Import log not found."}, status=status.HTTP_404_NOT_FOUND)
        log.delete()
        return Response({"message": "Import log deleted."}, status=status.HTTP_200_OK)
